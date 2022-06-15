import requests
import openpyxl
from datetime import date
import shutil
import sys

print("Python Executing..")
todays_date = date.today()
usnList = []
permittedSem = []

collegeCodeParam = sys.argv[1]
yearParam = sys.argv[2]
branchParam = sys.argv[3]
startUsnParam = sys.argv[4]
endUsnParam = sys.argv[5]
folderPathParam= sys.argv[6]

# collegeCodeParam = "4CB"
# yearParam = "19"
# branchParam = "IS"
# startUsnParam = "1"
# endUsnParam = "1"

for i in range(int(startUsnParam), int(endUsnParam)+1):
    if(i < 10):
        usnList.append(collegeCodeParam+yearParam+branchParam+"00"+str(i))
    elif(i < 100):
        usnList.append(collegeCodeParam+yearParam+branchParam+"0"+str(i))
    else:
        usnList.append(collegeCodeParam+yearParam+branchParam+str(i))
    startyear = int("20"+yearParam)
    firstyear = startyear+1
    secondyear = startyear+2
    thirdyear = startyear+3
    fourthyear = startyear+4
if(todays_date.year == firstyear):
    permittedSem.append(1)
    permittedSem.append(2)
elif(todays_date.year == secondyear):
    permittedSem.append(3)
    permittedSem.append(4)
elif(todays_date.year == thirdyear):
    permittedSem.append(5)
    permittedSem.append(6)
elif(todays_date.year == fourthyear):
    permittedSem.append(7)
    permittedSem.append(8)


shutil.copy('ExcelFSTemplate.xlsx', folderPathParam)

excelfile= folderPathParam+"ExcelFSTemplate.xlsx"
wb_obj = openpyxl.load_workbook(excelfile)
subcounter = 0

print("Generating...")
for sheet in wb_obj.sheetnames:
    rowCount = 5
    firstTime = True
    sheetName = sheet
    sheet_obj = wb_obj[sheetName]

    for usn in usnList:
        querryString = "https://api.vtuconnect.in/result/"+usn
        r = requests.get(querryString,
                         headers={
                             "Accept": "application/json",
                             "Authorization": "Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VyIjp7InNjb3BlIjpbInVzZXIiXSwiZW1haWwiOiJuYW5kYW5AZ21haWwuY29tIn0sImlhdCI6MTU0NzMxODI5N30.ZPO8tf03azhTJ1qmgSVyGV80k9EfomXgGazdLyUC6fw",
                             "Host": "api.vtuconnect.in",
                         })
        data = r.json()

        for i in data:
            if(int(i["resultMonthYear"].split()[1]) == todays_date.year and (i["semester"] == str(permittedSem[0]) or i["semester"] == str(permittedSem[1]))):
                if(firstTime):
                    examNameCell = sheet_obj.cell(row=1, column=1)
                    examNameCell.value = i["resultMonthYear"]
                    subNameCell = sheet_obj.cell(row=2, column=3)
                    subNameCell.value = i["subjects"][subcounter]["subjectName"]
                    subCodeCell = sheet_obj.cell(row=3, column=3)
                    subCodeCell.value = i["subjects"][subcounter]["subjectCode"]
                    firstTime = False
                nameCell = sheet_obj.cell(row=rowCount, column=1)
                usnCell = sheet_obj.cell(row=rowCount, column=2)
                nameCell.value = i["name"]
                usnCell.value = i["usn"]
                iamarksCell = sheet_obj.cell(row=rowCount, column=3)
                iamarksCell.value = i["subjects"][subcounter]["iaMarks"]
                emarksCell = sheet_obj.cell(row=rowCount, column=4)
                emarksCell.value = i["subjects"][subcounter]["eMarks"]
                totalmarksCell = sheet_obj.cell(row=rowCount, column=5)
                totalmarksCell.value = i["subjects"][subcounter]["total"]
                resultsCell = sheet_obj.cell(row=rowCount, column=6)
                resultsCell.value = i["subjects"][subcounter]["result"]
                rowCount = rowCount+1
                wb_obj.save("D:/ExcelFSTemplate.xlsx")
    subcounter = subcounter+1

print("Successfully Completed")
